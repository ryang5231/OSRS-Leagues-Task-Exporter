from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import numbers
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.datavalidation import DataValidation
import io
from openpyxl.utils import get_column_letter
import helper
from openpyxl.worksheet.errors import IgnoredError

IDX_TASK_DIFFICULTY = 0
IDX_VERBOSE_DESCRIPTION = 1
IDX_SKILL_REQUIREMENTS = 2
IDX_PERCENT_COMPL = 3
WRAP_COL_LETTERS = ["A", "B", "E"]
CUSTOM_COL_SETTINGS = {
                        "A": {"wrap": True, "col_width": 20},
                        "B": {"wrap": True, "col_width": 20},
                        "E": {"wrap": True, "col_width": 30},
                        "F": {"col_width": 15, "format": numbers.FORMAT_TEXT},
                        "G": {"col_width": 15},
                    }
TASKS_URL = 'https://oldschool.runescape.wiki/w/Shattered_Relics_League/Tasks'
TABLE_HEADERS = ["Task", "Description", "Difficulty", "Points", "Requirement(s)", "% Completed", "Completed?"]
FILE_NAME = "OSRS_3_Shattered_League_Tasks.xlsx"
SHEET_NAME = "Tasks"

def get_task_excel():

    response = helper.fetch_html(TASKS_URL)

    soup = BeautifulSoup(response.content, 'html.parser')

    # full div with point info and task descr
    points_reference = {
                        "Beginner": 5,
                        "Easy": 5,
                        "Medium": 25,
                        "Hard": 50,
                        "Elite": 125,
                        "Master": 250
                        }

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(TABLE_HEADERS)

    rows = soup.find_all("tr", attrs={"data-taskid": True})
    # point info: span title
    # text within td
    for r in rows:
        cols = r.find_all("td")
        if cols:
            task_difficulty = cols[IDX_TASK_DIFFICULTY].find("span", title=True)["title"]
            task_title = cols[IDX_TASK_DIFFICULTY].get_text(" ", strip=True)
            points = points_reference[task_difficulty]
            verbose = helper.text_cleaner(cols[IDX_VERBOSE_DESCRIPTION].get_text(" ", strip=False))
            skill_req = helper.text_cleaner(cols[IDX_SKILL_REQUIREMENTS].get_text(" ", strip=True))
            percent_compl = cols[IDX_PERCENT_COMPL].get_text(" ", strip=True)
            
            # if not '<' in percent_compl:
            #     percent_compl = percent_compl.replace("%", "")
            #     try:
            #         percent_compl = int(percent_compl)
            #     except ValueError:
            #         try:
            #             percent_compl = float(percent_compl)
            #         except:
            #             percent_compl = -1
            #             print("Invalid Value!")

            ws.append([task_title, verbose, task_difficulty, points, skill_req, percent_compl, False])

    num_rows = ws.max_row       # number of rows with data (includes header)
    num_cols = ws.max_column    # number of columns with data

    start_cell = "A1"
    end_cell = f"{get_column_letter(num_cols)}{num_rows}"
    table_ref = f"{start_cell}:{end_cell}"

    table = Table(displayName=SHEET_NAME, ref=table_ref)
    ws.add_table(table)

    ws = helper.format_columns(ws, CUSTOM_COL_SETTINGS)

    completed_col_letter = get_column_letter(num_cols)  # should be "G"
    dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    dv.add(f"{completed_col_letter}2:{completed_col_letter}{num_rows}")
    ws.add_data_validation(dv)

    wb.save(FILE_NAME)
    # print("file successfully saved. sending back...")
    # output = io.BytesIO()
    # wb.save(output)
    # output.seek(0)
    # return output

get_task_excel()
