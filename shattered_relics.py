import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table
import re
from openpyxl.utils import get_column_letter

TASK_DIFFICULTY_IDX = 0
VERBOSE_IDX = 1
SKILL_REQUIREMENT_IDX = 2
PERCENT_COMPL_IDX = 3
rows_to_wrap = ["A", "B", "E"]
PERCENTAGE_ROW_LETTER = "F"
TASKS_URL = 'https://oldschool.runescape.wiki/w/Shattered_Relics_League/Tasks'
REQUEST_HEADERS = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 \
                    (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Referer": "https://www.google.com/",
    }
TABLE_HEADERS = ["Task", "Description", "Difficulty", "Points", "Requirement(s)", "% Completed"]
FILE_NAME = "OSRS_Shattered_League_Tasks.xlsx"
SHEET_NAME = "Tasks"

def text_cleaner(input_text):
    result = input_text
    # Replace multiple spaces with one space
    result = re.sub(r' +', ' ', result)
    replacements = [
        (" .", "."),
        ("\n", ""),
        ("( ", "("),
        (" )", ")"),
        (" ,", ","),
        (" ;", ";"),
        ('[ ', '['),
        (' ]', ']')
    ]
    for old, new in replacements:
        result = result.replace(old, new)
    return result.strip()


def get_task_excel():
    response = requests.get(TASKS_URL, headers=REQUEST_HEADERS)

    # print(response.status_code)
    # print(response.content)

    soup = BeautifulSoup(response.content, 'html.parser')

    # full div with point info and task descr
    points_reference = {
                        "Beginner": 5,
                        "Easy": 5,
                        "Medium": 25,
                        "Hard": 50,
                        "Elite": 125,
                        "Master": 250
                        }

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(TABLE_HEADERS)

    rows = soup.find_all("tr", attrs={"data-taskid": True})
    # point info: span title
    # text within td
    for r in rows:
        cols = r.find_all("td")
        if cols:
            task_difficulty = cols[TASK_DIFFICULTY_IDX].find("span", title=True)["title"]
            task_title = cols[TASK_DIFFICULTY_IDX].get_text(" ", strip=True)
            points = points_reference[task_difficulty]
            verbose = text_cleaner(cols[VERBOSE_IDX].get_text(" ", strip=False))
            skill_req = text_cleaner(cols[SKILL_REQUIREMENT_IDX].get_text(" ", strip=True))
            percent_compl = cols[PERCENT_COMPL_IDX].get_text(" ", strip=True)

            if not '<' in percent_compl:
                percent_compl = percent_compl.replace("%", "")
                try:
                    percent_compl = int(percent_compl)
                except ValueError:
                    try:
                        percent_compl = float(percent_compl)
                    except:
                        percent_compl = -1
                        print("Invalid Value!")

            ws.append([task_title, verbose, task_difficulty, points, skill_req, percent_compl])

    num_rows = ws.max_row       # number of rows with data (includes header)
    num_cols = ws.max_column    # number of columns with data

    start_cell = "A1"
    end_cell = f"{get_column_letter(num_cols)}{num_rows}"
    table_ref = f"{start_cell}:{end_cell}"

    table = Table(displayName=SHEET_NAME, ref=table_ref)
    ws.add_table(table)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['E'].width = 30

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        if col_letter in rows_to_wrap:
            for cell in col:
                try:
                    if cell.value:
                        cell.alignment = Alignment(
                                wrap_text=True,
                                horizontal="left",
                                vertical="top"
                            )
                except:
                    pass
        if col_letter == PERCENTAGE_ROW_LETTER:
            for cell in col:
                cell.alignment = Alignment(
                                    horizontal="right",
                                    vertical="bottom"
                                )
                if isinstance(cell.value, int):
                    cell.value /= 100
                    cell.number_format = '0%'
                elif isinstance(cell.value, float):
                    cell.value /= 100
                    cell.number_format = '0.0%'


    wb.save(FILE_NAME)
