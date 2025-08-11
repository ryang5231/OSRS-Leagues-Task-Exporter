import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.errors import IgnoredErrors
from openpyxl.utils import get_column_letter

TASK_DIFFICULTY_IDX = 0
VERBOSE_IDX = 1
SKILL_REQUIREMENT_IDX = 2
PERCENT_COMPL_IDX = 3
rows_to_wrap = ["A", "B", "E"]
PERCENTAGE_ROW_LETTER = "F"

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 \
                   (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Referer": "https://www.google.com/",
}

url = 'https://oldschool.runescape.wiki/w/Shattered_Relics_League/Tasks'
response = requests.get(url, headers=headers)

# print(response.status_code)
# print(response.content)

soup = BeautifulSoup(response.content, 'html.parser')

# full div with point info and task descr
points_reference = {
                    "Beginner": 5,
                    "Easy": 5,
                    "Medium": 25,
                    "Hard": 50,
                    "Elite": 125,
                    "Master": 250
                    }

# data = {
#     "Task": [],
#     "Description": [],
#     "Difficulty": [],
#     "Points": [],
#     "Requirement(s)": [],
#     "% Completed": []
# }

wb = Workbook()
ws = wb.active
ws.title = "Tasks"
ws.append(["Task", "Description", "Difficulty", "Points", "Requirement(s)", "% Completed"])

rows = soup.find_all("tr", attrs={"data-taskid": True})
# point info: span title
# text within td
for r in rows:
    cols = r.find_all("td")
    if cols:
        task_difficulty = cols[TASK_DIFFICULTY_IDX].find("span", title=True)["title"]
        task_title = cols[TASK_DIFFICULTY_IDX].get_text(" ", strip=True)
        points = points_reference[task_difficulty]
        verbose = cols[VERBOSE_IDX].get_text(" ", strip=False)

        verbose = (
            verbose
            .replace("  ", " ")
            .replace(" .", ".")
            .replace("\n", "")
        )

        skill_req = cols[SKILL_REQUIREMENT_IDX].get_text(" ", strip=True)
        percent_compl = cols[PERCENT_COMPL_IDX].get_text(" ", strip=True)

        # data["Task"].append(task_title)
        # data["Description"].append(verbose)
        # data["Difficulty"].append(task_difficulty)
        # data["Points"].append(points)
        # data["Requirement(s)"].append(skill_req)
        # data["% Completed"].append(percent_compl)

        ws.append([task_title, verbose, task_difficulty, points, skill_req, percent_compl])

num_rows = ws.max_row       # number of rows with data (includes header)
num_cols = ws.max_column    # number of columns with data

# Get Excel-style column letters, e.g. A, B, C
start_cell = "A1"
end_cell = f"{get_column_letter(num_cols)}{num_rows}"
table_ref = f"{start_cell}:{end_cell}"

table = Table(displayName="Tasks", ref=table_ref)
ws.add_table(table)

# df = pd.DataFrame(data)
file_name = "OSRS_League_Tasks.xlsx"
# df.to_excel(file_name, index=False)

# wb = load_workbook(file_name)
# ws = wb.active

for col in ws.columns:
    max_length = 0
    col_letter = col[0].column_letter
    if col_letter in rows_to_wrap:
        ws.column_dimensions[col_letter].width = 30
        for cell in col:
            try:
                if cell.value:
                    cell.alignment = Alignment(
                            wrap_text=True,
                            horizontal=cell.alignment.horizontal or "left",
                            vertical=cell.alignment.vertical or "top"
                        )
            except:
                pass


wb.save(file_name)

# print(data)
# print(soup.prettify())

